"""L10 D3 / L11 D4 — Excel/CSV data ingestion + ForecastInputs builder.

Bridges the web form (8 numerical inputs + 3 file uploads) to the L6-H
``ForecastInputs`` dataclass consumed by ``aggregate_ensemble``.

Design notes
------------
* **Excel parsing** uses ``openpyxl`` for ``.xlsx`` and the stdlib ``csv`` module
  for ``.csv``. The "data" sheet is preferred when present; otherwise the
  active sheet is used.
* **Validation is strict**: missing required columns, non-finite numerics, empty
  files, and missing date columns all produce a structured ``IngestionResult``
  with ``success=False`` and a Vietnamese-friendly ``error`` string.
* **ForecastInputs derivation (L11)**: the primary path delegates to
  ``macro_pipeline.webapp.producer_adapter.ProducerAdapter``, which derives all
  six ForecastInputs fields from the bundled L11 data snapshot
  (`macro_pipeline/data_snapshot/`) with the form values overlaid as the latest
  observation. If the snapshot is missing or the adapter raises for any reason,
  the call transparently falls back to the legacy L10 heuristic-modulator path
  (preserved verbatim as ``_build_heuristic``). The provenance is recorded on
  ``ForecastInputsBuilder.last_provenance`` for UI display.

Public API
----------
``IngestionResult``       Frozen result wrapper (success / data / error).
``ExcelDataIngester``     Parses yield curve / credit spreads / sentiment files.
``ForecastInputsBuilder`` Maps parsed data + numerical form inputs to ``ForecastInputs``.
"""
from __future__ import annotations

import csv
import logging
import math
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from macro_pipeline.ensemble.aggregator import SUPPORTED_HORIZONS, ForecastInputs

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Public result wrapper
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class IngestionResult:
    """Outcome of a single Excel/CSV parse call."""

    success: bool
    data: dict[str, Any] | None = None
    error: str | None = None


# ---------------------------------------------------------------------------
# Excel / CSV reader helpers
# ---------------------------------------------------------------------------
def _load_rows(path: Path) -> list[dict[str, Any]]:
    """Load rows from an .xlsx (sheet "data" if present, else active) or .csv.

    Returns a list of dicts keyed by lower-cased column name. Raises
    ``ValueError`` on unsupported extension or empty file.
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            rows = [
                {(k or "").strip().lower(): v for k, v in row.items()}
                for row in reader
            ]
    elif suffix == ".xlsx":
        try:
            wb = load_workbook(
                filename=str(path), data_only=True, read_only=True
            )
        except Exception as exc:
            raise ValueError(
                f"Không đọc được file Excel ({type(exc).__name__}): {exc}. "
                "Kiểm tra file có đúng định dạng .xlsx và không bị hỏng."
            ) from exc
        try:
            sheet_name = "data" if "data" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet_name]
            iterator = ws.iter_rows(values_only=True)
            try:
                header = [
                    (str(c).strip().lower() if c is not None else "")
                    for c in next(iterator)
                ]
            except StopIteration as exc:
                raise ValueError(
                    "File trống (không có hàng tiêu đề)."
                ) from exc
            rows = []
            for raw_row in iterator:
                if raw_row is None:
                    continue
                if all(c is None or (isinstance(c, str) and not c.strip()) for c in raw_row):
                    continue
                rows.append({header[i]: raw_row[i] for i in range(len(header))})
        finally:
            wb.close()
    else:
        raise ValueError(
            f"Định dạng file không hỗ trợ: {suffix!r}. Chỉ chấp nhận .xlsx hoặc .csv."
        )
    if not rows:
        raise ValueError("File không có dữ liệu (sau hàng tiêu đề).")
    return rows


def _to_finite_float(raw: Any, column: str, row_idx: int) -> float:
    """Coerce a cell to a finite float; raise ``ValueError`` with a useful message."""
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        raise ValueError(
            f"Cột '{column}' hàng {row_idx}: giá trị trống."
        )
    try:
        value = float(raw)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Cột '{column}' hàng {row_idx}: '{raw}' không phải là số."
        ) from exc
    if not math.isfinite(value):
        raise ValueError(
            f"Cột '{column}' hàng {row_idx}: giá trị {value} không hữu hạn."
        )
    return value


def _require_columns(rows: list[dict[str, Any]], required: tuple[str, ...]) -> None:
    """Raise ``ValueError`` if any required column is missing from the header."""
    first = rows[0]
    missing = [c for c in required if c not in first]
    if missing:
        raise ValueError(
            f"Thiếu cột bắt buộc: {missing}. Cột có sẵn: {list(first.keys())}."
        )


# ---------------------------------------------------------------------------
# Excel ingester
# ---------------------------------------------------------------------------
class ExcelDataIngester:
    """Parse the three L10 Excel/CSV upload types into structured dicts.

    All methods accept either ``.xlsx`` or ``.csv`` paths and return an
    ``IngestionResult``. Validation is strict — required columns must be
    present, all numeric cells must be finite, and at least one data row must
    exist.
    """

    YIELD_CURVE_REQUIRED = ("date", "2y", "10y")
    CREDIT_SPREADS_REQUIRED = ("date", "ig_oas")
    SENTIMENT_REQUIRED = ("date", "indicator", "value")

    def parse_yield_curve(self, path: Path) -> IngestionResult:
        """Parse yield-curve file. Required columns: date, 2y, 10y.

        Returns ``data`` dict containing ``rows`` (parsed list), ``latest`` (last
        row), and ``inverted`` (bool — True if 2Y > 10Y at the latest date).
        """
        try:
            rows = _load_rows(path)
            _require_columns(rows, self.YIELD_CURVE_REQUIRED)
            parsed: list[dict[str, Any]] = []
            for idx, row in enumerate(rows, start=2):
                two_y = _to_finite_float(row["2y"], "2y", idx)
                ten_y = _to_finite_float(row["10y"], "10y", idx)
                parsed.append({"date": row["date"], "2y": two_y, "10y": ten_y})
            latest = parsed[-1]
            inverted = latest["2y"] > latest["10y"]
        except ValueError as exc:
            return IngestionResult(success=False, error=str(exc))
        return IngestionResult(
            success=True,
            data={"rows": parsed, "latest": latest, "inverted": inverted},
        )

    def parse_credit_spreads(self, path: Path) -> IngestionResult:
        """Parse credit-spreads file. Required columns: date, ig_oas.

        Returns ``data`` dict containing ``rows``, ``latest``, and ``elevated``
        (bool — True if latest IG OAS > 200 bps, a common stress threshold).
        """
        try:
            rows = _load_rows(path)
            _require_columns(rows, self.CREDIT_SPREADS_REQUIRED)
            parsed: list[dict[str, Any]] = []
            for idx, row in enumerate(rows, start=2):
                ig_oas = _to_finite_float(row["ig_oas"], "ig_oas", idx)
                parsed.append({"date": row["date"], "ig_oas": ig_oas})
            latest = parsed[-1]
            elevated = latest["ig_oas"] > 200.0
        except ValueError as exc:
            return IngestionResult(success=False, error=str(exc))
        return IngestionResult(
            success=True,
            data={"rows": parsed, "latest": latest, "elevated": elevated},
        )

    def parse_sentiment(self, path: Path) -> IngestionResult:
        """Parse sentiment file. Required columns: date, indicator, value.

        Returns ``data`` dict containing ``rows``, ``latest``, and ``by_indicator``
        (dict mapping indicator name to its most recent value).
        """
        try:
            rows = _load_rows(path)
            _require_columns(rows, self.SENTIMENT_REQUIRED)
            parsed: list[dict[str, Any]] = []
            for idx, row in enumerate(rows, start=2):
                value = _to_finite_float(row["value"], "value", idx)
                parsed.append(
                    {
                        "date": row["date"],
                        "indicator": str(row["indicator"]).strip(),
                        "value": value,
                    }
                )
            by_indicator: dict[str, float] = {}
            for item in parsed:
                by_indicator[item["indicator"]] = item["value"]
            latest = parsed[-1]
        except ValueError as exc:
            return IngestionResult(success=False, error=str(exc))
        return IngestionResult(
            success=True,
            data={
                "rows": parsed,
                "latest": latest,
                "by_indicator": by_indicator,
            },
        )


# ---------------------------------------------------------------------------
# ForecastInputs builder
# ---------------------------------------------------------------------------

# Canonical defaults — mirror tests/test_aggregator.py helper (L6-H proven).
_DEFAULTS_POINT_ESTIMATES: dict[int, float] = {1: 0.05, 3: 0.06, 5: 0.065, 10: 0.07}
_DEFAULTS_N_EFF: dict[int, int] = {1: 100, 3: 30, 5: 18, 10: 9}
_DEFAULTS_FORECAST_SIGMAS: dict[int, float] = {1: 0.02, 3: 0.025, 5: 0.03, 10: 0.035}
_DEFAULTS_ANALOG_DISPERSIONS: dict[int, float] = {1: 0.04, 3: 0.05, 5: 0.06, 10: 0.07}
_DEFAULTS_RETURN_SIGMAS: dict[int, float] = {1: 0.15, 3: 0.16, 5: 0.17, 10: 0.18}
_DEFAULTS_RECESSION_P: dict[int, float] = {1: 0.15, 3: 0.25, 5: 0.35, 10: 0.45}


@dataclass(frozen=True)
class BuildResult:
    """Result of ``ForecastInputsBuilder.build``."""

    inputs: ForecastInputs
    defaults_used: tuple[str, ...] = field(default_factory=tuple)


class ForecastInputsBuilder:
    """Map web form + uploaded data to a valid ``ForecastInputs``.

    L11 primary path: delegate to ``ProducerAdapter``, which derives all six
    ForecastInputs fields from real historical panels in the bundled L11 data
    snapshot, with the form values overlaid as the latest observation. The
    panel-derived path produces honest cross-period dispersion + realized vol +
    NBER-base-rate recession probabilities; previously L10 used heuristic
    modulators around canonical L6-H defaults.

    L11 fallback path (preserved verbatim from L10): if the snapshot is missing
    or the adapter raises, ``_build_heuristic`` runs the original heuristic
    modulator block. This guarantees the web app keeps working even when V's
    snapshot is unbuilt.

    Public API (BINDING — Gate 8): ``build()`` signature unchanged from L10.
    L11 adds ``last_provenance`` instance attribute set after every ``build()``
    call (route handlers read it for results.html provenance display).
    """

    LONG_RUN_REAL_RETURN = 0.065

    def __init__(self, producer_adapter: Any | None = None) -> None:
        # ``producer_adapter`` typed as ``Any`` to avoid hard import of
        # ProducerAdapter at class-definition time (snapshot_loader must remain
        # importable in environments where pandas is partially installed).
        self._adapter = producer_adapter
        self.last_provenance: dict[str, Any] = {}

    def _get_adapter(self):
        if self._adapter is None:
            from macro_pipeline.webapp.producer_adapter import ProducerAdapter
            self._adapter = ProducerAdapter()
        return self._adapter

    def build(
        self,
        uploaded_data: dict[str, dict[str, Any]] | None,
        numerical_inputs: dict[str, float],
        horizons: tuple[int, ...] = SUPPORTED_HORIZONS,
    ) -> ForecastInputs:
        """Build a valid ForecastInputs (panel-derived primary, heuristic fallback).

        Parameters
        ----------
        uploaded_data
            Mapping with optional keys ``yield_curve``, ``credit_spreads``,
            ``sentiment`` (each mapping to the ``data`` dict returned by
            ``ExcelDataIngester``). May be empty.
        numerical_inputs
            Mapping with the 8 form fields (see ``REQUIRED_NUMERICAL_FIELDS``).
        horizons
            Horizons the result must cover (always equals SUPPORTED_HORIZONS
            internally; param exists for forward compat).

        Returns
        -------
        ForecastInputs
            Validated immutable inputs ready for ``aggregate_ensemble``.

        Side effects
        ------------
        Sets ``self.last_provenance`` to a dict describing which path ran +
        which panels/producers + any fallbacks.

        Raises
        ------
        ValueError
            If a required numerical input is missing or non-finite.
        """
        uploaded_data = uploaded_data or {}
        self._require_numerical(numerical_inputs)

        # ---- L11 primary path: producer-derived from bundled snapshot ----
        try:
            from macro_pipeline.webapp.snapshot_loader import SnapshotNotFoundError

            adapter = self._get_adapter()
            result = adapter.derive_forecast_inputs(numerical_inputs, uploaded_data)
            self.last_provenance = dict(result.provenance)
            return result.inputs
        except SnapshotNotFoundError as exc:
            self.last_provenance = self._fallback_provenance(
                f"snapshot_missing: {exc}"
            )
        except Exception as exc:
            log.exception("ProducerAdapter failed; falling back to heuristic")
            self.last_provenance = self._fallback_provenance(
                f"{type(exc).__name__}: {exc}"
            )

        # ---- L10 fallback path: heuristic modulators (preserved verbatim) --
        return self._build_heuristic(uploaded_data, numerical_inputs)

    def _fallback_provenance(self, reason: str) -> dict[str, Any]:
        return {
            "mode": "heuristic_fallback",
            "fallback_reason": reason,
            "snapshot_date": "n/a",
            "panels_used": (),
            "producers_run": (),
            "fallbacks": (("producer_chain", reason),),
            "form_overlay_applied": True,
        }

    def _build_heuristic(
        self,
        uploaded_data: dict[str, dict[str, Any]],
        numerical_inputs: dict[str, float],
    ) -> ForecastInputs:
        """L10 heuristic-modulator path (preserved verbatim as L11 fallback)."""
        pmi_avg = 0.5 * (
            numerical_inputs["pmi_manufacturing"]
            + numerical_inputs["pmi_services"]
        )
        cape = numerical_inputs["cape_ratio"]
        unemployment = numerical_inputs["unemployment_rate"]

        yield_curve = uploaded_data.get("yield_curve") or {}
        yc_inverted = bool(yield_curve.get("inverted", False))

        point_estimates: dict[int, float] = {}
        recession_probabilities: dict[int, float] = {}
        for horizon in SUPPORTED_HORIZONS:
            point_estimates[horizon] = self._modulated_point_estimate(
                horizon, pmi_avg, cape, unemployment
            )
            recession_probabilities[horizon] = self._modulated_recession_p(
                horizon, pmi_avg, unemployment, yc_inverted
            )

        return ForecastInputs(
            point_estimates=point_estimates,
            point_estimate_n_eff=dict(_DEFAULTS_N_EFF),
            forecast_sigmas=dict(_DEFAULTS_FORECAST_SIGMAS),
            analog_dispersions=dict(_DEFAULTS_ANALOG_DISPERSIONS),
            return_sigmas=dict(_DEFAULTS_RETURN_SIGMAS),
            recession_probabilities=recession_probabilities,
        )

    # ---- internals ------------------------------------------------------
    REQUIRED_NUMERICAL_FIELDS: tuple[str, ...] = (
        "pmi_manufacturing",
        "pmi_services",
        "cape_ratio",
        "sp500_current",
        "payrolls_mom",
        "unemployment_rate",
        "core_cpi_yoy",
        "fed_funds_rate",
    )

    def _require_numerical(self, numerical_inputs: dict[str, float]) -> None:
        missing = [
            f for f in self.REQUIRED_NUMERICAL_FIELDS if f not in numerical_inputs
        ]
        if missing:
            raise ValueError(
                f"Thiếu trường số bắt buộc: {missing}"
            )
        for field_name in self.REQUIRED_NUMERICAL_FIELDS:
            value = numerical_inputs[field_name]
            if not isinstance(value, (int, float)):
                raise ValueError(
                    f"Trường '{field_name}' phải là số (nhận được "
                    f"{type(value).__name__})."
                )
            if not math.isfinite(float(value)):
                raise ValueError(
                    f"Trường '{field_name}' không hữu hạn: {value!r}."
                )

    def _modulated_point_estimate(
        self,
        horizon: int,
        pmi_avg: float,
        cape: float,
        unemployment: float,
    ) -> float:
        """Heuristic point estimate by horizon. Bounded to [-0.10, 0.20]."""
        base = _DEFAULTS_POINT_ESTIMATES[horizon]
        if horizon == 10:
            cape_adjust = -0.001 * (cape - 22.0)
            return _clip(base + cape_adjust, -0.10, 0.20)
        pmi_signal = (pmi_avg - 50.0) / 100.0
        unemployment_signal = (5.0 - unemployment) / 100.0
        cyclical = 0.05 * pmi_signal + 0.10 * unemployment_signal
        weight_cyclical = {1: 0.7, 3: 0.5, 5: 0.3}[horizon]
        modulated = (
            weight_cyclical * cyclical
            + (1.0 - weight_cyclical) * base
        )
        return _clip(modulated, -0.10, 0.20)

    def _modulated_recession_p(
        self,
        horizon: int,
        pmi_avg: float,
        unemployment: float,
        yc_inverted: bool,
    ) -> float:
        """Heuristic recession probability. Bounded to [0.02, 0.95]."""
        base = _DEFAULTS_RECESSION_P[horizon]
        bump = 0.0
        if pmi_avg < 45.0:
            bump += 0.20
        elif pmi_avg < 50.0:
            bump += 0.10
        if unemployment > 5.5:
            bump += 0.15
        if yc_inverted:
            bump += 0.20
        horizon_multiplier = {1: 1.0, 3: 1.05, 5: 1.10, 10: 1.15}[horizon]
        modulated = (base + bump) * horizon_multiplier
        return _clip(modulated, 0.02, min(0.95, base * 2.0 + 0.20))


def _clip(value: float, low: float, high: float) -> float:
    """Clip to ``[low, high]`` while preserving finite invariant."""
    if not math.isfinite(value):
        raise ValueError(f"Cannot clip non-finite value: {value!r}")
    return max(low, min(high, value))
