"""L10 D6 — Generate three Excel templates for V's web app input.

Per Strategic L10 single comprehensive pre-flight 2026-05-17.

Each template has two sheets:
* ``instructions`` (Vietnamese) — purpose, required columns, data sources,
  minimum row count.
* ``data`` — header row + one sample row (V deletes the sample + adds real data).

Output to ``macro_pipeline/webapp/static/templates/`` so the Flask
``/forecast/template/<name>`` endpoint can ``send_from_directory`` them.

Usage
-----
    python scripts/generate_excel_templates.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
SAMPLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _write_instructions_sheet(
    wb: Workbook,
    title: str,
    purpose: str,
    columns: list[tuple[str, str]],
    sources: list[str],
    min_rows: int,
) -> None:
    """Write a standard Vietnamese instructions sheet."""
    ws = wb.create_sheet("instructions", index=0)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:D1")

    row = 3
    ws.cell(row=row, column=1, value="Mục đích:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=purpose).alignment = Alignment(wrap_text=True)
    row += 2

    ws.cell(row=row, column=1, value="Cột bắt buộc:").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Tên cột").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="Mô tả").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    row += 1
    for col_name, col_desc in columns:
        ws.cell(row=row, column=1, value=col_name)
        ws.cell(row=row, column=2, value=col_desc).alignment = Alignment(wrap_text=True)
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Nguồn dữ liệu khuyến nghị:").font = Font(bold=True)
    row += 1
    for src in sources:
        ws.cell(row=row, column=1, value=f"• {src}").alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
    row += 1

    ws.cell(row=row, column=1, value=f"Số hàng tối thiểu: {min_rows}").font = Font(
        italic=True, color="C00000"
    )
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=(
            "Lưu ý: sheet 'data' bên dưới có 1 hàng mẫu (highlight vàng). "
            "Xóa hàng mẫu và điền dữ liệu thật vào, sau đó upload file lên web app."
        ),
    ).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18


def _write_data_sheet(
    wb: Workbook,
    headers: list[str],
    sample_row: list,
) -> None:
    """Write the ``data`` sheet — headers + one highlighted sample row."""
    ws = wb.create_sheet("data")
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for idx, value in enumerate(sample_row, start=1):
        cell = ws.cell(row=2, column=idx, value=value)
        cell.fill = SAMPLE_FILL
    for col_letter in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = 18


# ---------------------------------------------------------------------------
# Template generators
# ---------------------------------------------------------------------------
def generate_yield_curve(output_dir: Path) -> Path:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    _write_instructions_sheet(
        wb,
        title="YIELD CURVE TEMPLATE — Macro Forecast Terminal",
        purpose=(
            "Cung cấp dữ liệu lãi suất Treasury 2 năm và 10 năm theo thời gian. "
            "Web app sử dụng để phát hiện yield curve inversion (2Y > 10Y) — "
            "một tín hiệu suy thoái cổ điển."
        ),
        columns=[
            ("date", "Ngày quan sát (YYYY-MM-DD)"),
            ("2y", "Lãi suất Treasury 2 năm (%, ví dụ 4.85)"),
            ("10y", "Lãi suất Treasury 10 năm (%, ví dụ 4.30)"),
        ],
        sources=[
            "FRED: DGS2 (2-Year Treasury Constant Maturity)",
            "FRED: DGS10 (10-Year Treasury Constant Maturity)",
            "TradingView: US02Y và US10Y",
            "U.S. Treasury Daily Yield Curve Rates",
        ],
        min_rows=12,
    )
    _write_data_sheet(
        wb,
        headers=["date", "2y", "10y"],
        sample_row=["2026-04-30", 4.85, 4.30],
    )
    output_path = output_dir / "yield-curve.xlsx"
    wb.save(str(output_path))
    return output_path


def generate_credit_spreads(output_dir: Path) -> Path:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    _write_instructions_sheet(
        wb,
        title="CREDIT SPREADS TEMPLATE — Macro Forecast Terminal",
        purpose=(
            "Cung cấp dữ liệu Investment-Grade OAS (Option-Adjusted Spread) theo "
            "thời gian. Web app sử dụng để đánh giá credit stress — trên 200 bps "
            "thường báo hiệu stress, trên 400 bps là khủng hoảng."
        ),
        columns=[
            ("date", "Ngày quan sát (YYYY-MM-DD)"),
            ("ig_oas", "Investment-Grade OAS (basis points, ví dụ 110)"),
        ],
        sources=[
            "FRED: BAMLC0A0CM (ICE BofA US Corporate Index OAS)",
            "FRED: BAMLH0A0HYM2 (ICE BofA US HY Index OAS) — tùy chọn",
            "Bloomberg / WSJ Markets — Credit Spread tables",
        ],
        min_rows=12,
    )
    _write_data_sheet(
        wb,
        headers=["date", "ig_oas"],
        sample_row=["2026-04-30", 110.0],
    )
    output_path = output_dir / "credit-spreads.xlsx"
    wb.save(str(output_path))
    return output_path


def generate_sentiment(output_dir: Path) -> Path:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    _write_instructions_sheet(
        wb,
        title="SENTIMENT TEMPLATE — Macro Forecast Terminal",
        purpose=(
            "Cung cấp các chỉ số sentiment từ nhiều nguồn (AAII, NAAIM, ISM, ...). "
            "Mỗi hàng là một quan sát: date / tên indicator / giá trị. Web app sẽ "
            "lấy giá trị mới nhất của mỗi indicator để cập nhật state."
        ),
        columns=[
            ("date", "Ngày quan sát (YYYY-MM-DD)"),
            ("indicator", "Tên chỉ báo (ví dụ aaii_bull, naaim, ism_manuf)"),
            ("value", "Giá trị (số thập phân, ví dụ 35.0 cho 35%)"),
        ],
        sources=[
            "AAII Sentiment Survey — aaii.com/sentimentsurvey",
            "NAAIM Exposure Index — naaim.org/programs/naaim-exposure-index",
            "ISM Manufacturing/Services — ismworld.org",
            "VIX — CBOE / FRED VIXCLS",
        ],
        min_rows=12,
    )
    _write_data_sheet(
        wb,
        headers=["date", "indicator", "value"],
        sample_row=["2026-04-30", "aaii_bull", 35.2],
    )
    output_path = output_dir / "sentiment.xlsx"
    wb.save(str(output_path))
    return output_path


def main() -> None:
    repo_root = Path(__file__).parent.parent
    output_dir = (
        repo_root / "macro_pipeline" / "webapp" / "static" / "templates"
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    yc = generate_yield_curve(output_dir)
    cs = generate_credit_spreads(output_dir)
    se = generate_sentiment(output_dir)
    print(f"[OK] {yc.name} ({yc.stat().st_size:,} bytes)")
    print(f"[OK] {cs.name} ({cs.stat().st_size:,} bytes)")
    print(f"[OK] {se.name} ({se.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
