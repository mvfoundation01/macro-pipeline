"""L10 D12 — Tests for ``macro_pipeline.webapp.data_ingestion``.

PD18 strict ≥40% NEG: validation-heavy work; NEG-anchored by design.
Counts: 18 tests (10 NEG / 8 POS) = 55% NEG.
"""
from __future__ import annotations

import csv
from pathlib import Path

import pytest
from openpyxl import Workbook

from macro_pipeline.webapp.data_ingestion import (
    ExcelDataIngester,
    ForecastInputsBuilder,
    IngestionResult,
    _clip,
)


# ----------------------------------------------------------------------
# Fixture helpers
# ----------------------------------------------------------------------
def _write_xlsx(path: Path, headers: list[str], rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(str(path))


def _write_csv(path: Path, headers: list[str], rows: list[list]) -> None:
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)


def _good_numerical() -> dict[str, float]:
    return dict(
        pmi_manufacturing=51.0,
        pmi_services=53.0,
        cape_ratio=30.0,
        sp500_current=5800.0,
        payrolls_mom=180.0,
        unemployment_rate=4.1,
        core_cpi_yoy=2.9,
        fed_funds_rate=4.25,
    )


# ----------------------------------------------------------------------
# POS — happy paths (8 tests)
# ----------------------------------------------------------------------
def test_parse_yield_curve_xlsx_success(tmp_path: Path) -> None:
    p = tmp_path / "yc.xlsx"
    _write_xlsx(
        p,
        ["date", "2y", "10y"],
        [["2026-01-01", 4.85, 4.30], ["2026-02-01", 4.80, 4.20]],
    )
    res = ExcelDataIngester().parse_yield_curve(p)
    assert res.success is True
    assert res.data is not None
    assert res.data["inverted"] is True  # 2Y(4.80) > 10Y(4.20) at latest
    assert len(res.data["rows"]) == 2


def test_parse_yield_curve_csv_success(tmp_path: Path) -> None:
    p = tmp_path / "yc.csv"
    _write_csv(
        p,
        ["date", "2y", "10y"],
        [["2026-01-01", "3.0", "4.0"], ["2026-02-01", "3.1", "4.2"]],
    )
    res = ExcelDataIngester().parse_yield_curve(p)
    assert res.success is True
    assert res.data["inverted"] is False  # 2Y < 10Y


def test_parse_credit_spreads_success(tmp_path: Path) -> None:
    p = tmp_path / "cs.xlsx"
    _write_xlsx(
        p,
        ["date", "ig_oas"],
        [["2026-01-01", 95.0], ["2026-02-01", 250.0]],
    )
    res = ExcelDataIngester().parse_credit_spreads(p)
    assert res.success is True
    assert res.data["elevated"] is True  # 250 > 200


def test_parse_sentiment_success_by_indicator_aggregation(tmp_path: Path) -> None:
    p = tmp_path / "se.xlsx"
    _write_xlsx(
        p,
        ["date", "indicator", "value"],
        [
            ["2026-01-01", "aaii_bull", 33.0],
            ["2026-01-15", "naaim", 65.0],
            ["2026-02-01", "aaii_bull", 42.0],  # latest aaii_bull wins
        ],
    )
    res = ExcelDataIngester().parse_sentiment(p)
    assert res.success is True
    assert res.data["by_indicator"]["aaii_bull"] == pytest.approx(42.0)
    assert res.data["by_indicator"]["naaim"] == pytest.approx(65.0)


def test_builder_produces_all_4_horizons() -> None:
    b = ForecastInputsBuilder()
    fi = b.build(uploaded_data=None, numerical_inputs=_good_numerical())
    assert set(fi.point_estimates.keys()) == {1, 3, 5, 10}
    assert set(fi.recession_probabilities.keys()) == {1, 3, 5, 10}


def test_builder_pmi_above_50_lifts_short_horizon_estimate() -> None:
    b = ForecastInputsBuilder()
    weak = _good_numerical()
    weak.update(pmi_manufacturing=42.0, pmi_services=44.0)
    strong = _good_numerical()
    strong.update(pmi_manufacturing=58.0, pmi_services=56.0)
    fi_weak = b.build(uploaded_data=None, numerical_inputs=weak)
    fi_strong = b.build(uploaded_data=None, numerical_inputs=strong)
    assert fi_strong.point_estimates[1] > fi_weak.point_estimates[1]


def test_builder_yc_inversion_lifts_recession_p() -> None:
    b = ForecastInputsBuilder()
    base = _good_numerical()
    fi_no_inv = b.build(
        uploaded_data={"yield_curve": {"inverted": False}},
        numerical_inputs=base,
    )
    fi_inv = b.build(
        uploaded_data={"yield_curve": {"inverted": True}},
        numerical_inputs=base,
    )
    assert fi_inv.recession_probabilities[3] > fi_no_inv.recession_probabilities[3]


def test_clip_helper_bounds_values() -> None:
    assert _clip(0.05, 0.0, 1.0) == pytest.approx(0.05)
    assert _clip(-1.0, 0.0, 1.0) == 0.0
    assert _clip(2.0, 0.0, 1.0) == 1.0


# ----------------------------------------------------------------------
# NEG — strict validation (10 tests)
# ----------------------------------------------------------------------
def test_rejects_missing_required_column(tmp_path: Path) -> None:
    p = tmp_path / "yc.xlsx"
    _write_xlsx(p, ["date", "2y"], [["2026-01-01", 4.5]])  # missing 10y
    res = ExcelDataIngester().parse_yield_curve(p)
    assert res.success is False
    assert "10y" in (res.error or "")
    assert isinstance(res, IngestionResult)


def test_rejects_non_finite_numeric_value(tmp_path: Path) -> None:
    p = tmp_path / "cs.xlsx"
    _write_xlsx(
        p,
        ["date", "ig_oas"],
        [["2026-01-01", "not-a-number"]],
    )
    res = ExcelDataIngester().parse_credit_spreads(p)
    assert res.success is False
    assert res.error is not None
    assert "không phải là số" in res.error


def test_rejects_empty_file_no_data_rows(tmp_path: Path) -> None:
    p = tmp_path / "cs.xlsx"
    _write_xlsx(p, ["date", "ig_oas"], [])  # header only
    res = ExcelDataIngester().parse_credit_spreads(p)
    assert res.success is False
    assert res.error is not None
    assert "không có dữ liệu" in res.error.lower() or "trống" in res.error.lower()


def test_rejects_unsupported_extension(tmp_path: Path) -> None:
    p = tmp_path / "data.json"
    p.write_text('{"key":"value"}', encoding="utf-8")
    res = ExcelDataIngester().parse_yield_curve(p)
    assert res.success is False
    assert ".json" in (res.error or "") or "không hỗ trợ" in (res.error or "")


def test_rejects_csv_with_empty_cell(tmp_path: Path) -> None:
    p = tmp_path / "yc.csv"
    _write_csv(
        p,
        ["date", "2y", "10y"],
        [["2026-01-01", "", "4.3"]],  # empty 2y
    )
    res = ExcelDataIngester().parse_yield_curve(p)
    assert res.success is False
    assert "trống" in (res.error or "").lower()


def test_rejects_xlsx_with_inf_value(tmp_path: Path) -> None:
    p = tmp_path / "cs.xlsx"
    _write_xlsx(
        p,
        ["date", "ig_oas"],
        [["2026-01-01", float("inf")]],
    )
    res = ExcelDataIngester().parse_credit_spreads(p)
    assert res.success is False
    assert res.error is not None


def test_builder_rejects_missing_numerical_field() -> None:
    b = ForecastInputsBuilder()
    incomplete = _good_numerical()
    del incomplete["cape_ratio"]
    with pytest.raises(ValueError, match="cape_ratio"):
        b.build(uploaded_data=None, numerical_inputs=incomplete)


def test_builder_rejects_non_finite_numerical() -> None:
    b = ForecastInputsBuilder()
    bad = _good_numerical()
    bad["cape_ratio"] = float("nan")
    with pytest.raises(ValueError, match="cape_ratio"):
        b.build(uploaded_data=None, numerical_inputs=bad)


def test_builder_rejects_non_numeric_value() -> None:
    b = ForecastInputsBuilder()
    bad = _good_numerical()
    bad["cape_ratio"] = "thirty"  # type: ignore[assignment]
    with pytest.raises(ValueError, match="cape_ratio"):
        b.build(uploaded_data=None, numerical_inputs=bad)


def test_clip_rejects_non_finite_input() -> None:
    with pytest.raises(ValueError):
        _clip(float("nan"), 0.0, 1.0)
